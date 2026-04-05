"""
Utility helpers — PDF invoice & Excel export.
"""
import io
from decimal import Decimal
from datetime import datetime

from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ======================================================================
# PDF INVOICE
# ======================================================================
def generate_invoice_pdf(sale):
    """Generate a PDF invoice for a sale and return an HttpResponse."""

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=20 * mm, bottomMargin=20 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm)
    styles = getSampleStyleSheet()
    elements = []

    # -- Header --
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                  fontSize=20, alignment=TA_CENTER,
                                  spaceAfter=6)
    elements.append(Paragraph('Manamperi Rice Mill', title_style))

    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                alignment=TA_CENTER, fontSize=10,
                                spaceAfter=14)
    elements.append(Paragraph('Invoice / Bill', sub_style))

    # -- Invoice meta --
    meta_data = [
        ['Invoice #:', sale.invoice_number,
         'Date:', sale.created_at.strftime('%Y-%m-%d %H:%M')],
        ['Customer:', sale.customer.name if sale.customer else 'Walk-in',
         'Cashier:', sale.user.username if sale.user else '-'],
    ]
    meta_table = Table(meta_data, colWidths=[70, 160, 70, 160])
    meta_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 10))

    # -- Items table --
    header = ['#', 'Product', 'Qty (kg)', 'Unit Price', 'Subtotal']
    table_data = [header]
    for idx, item in enumerate(sale.items.select_related('product'), 1):
        table_data.append([
            str(idx),
            item.product.name,
            f"{item.quantity_kg:.2f}",
            f"Rs. {item.unit_price:.2f}",
            f"Rs. {item.subtotal:.2f}",
        ])

    items_table = Table(table_data, colWidths=[30, 180, 70, 80, 80])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#f5f5f5')]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 10))

    # -- Totals --
    right_style = ParagraphStyle('right', parent=styles['Normal'],
                                  alignment=TA_RIGHT, fontSize=10)
    totals_data = [
        ['', '', 'Total:', f"Rs. {sale.total_amount:.2f}"],
        ['', '', 'Discount:', f"Rs. {sale.discount:.2f}"],
        ['', '', 'Net Amount:', f"Rs. {sale.net_amount:.2f}"],
        ['', '', 'Cash Received:', f"Rs. {sale.cash_received:.2f}"],
        ['', '', 'Balance:', f"Rs. {sale.balance:.2f}"],
    ]
    totals_table = Table(totals_data, colWidths=[130, 100, 100, 100])
    totals_table.setStyle(TableStyle([
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('LINEABOVE', (2, 2), (-1, 2), 1, colors.black),
        ('LINEABOVE', (2, 4), (-1, 4), 1, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(totals_table)
    elements.append(Spacer(1, 20))

    # -- Footer --
    footer_style = ParagraphStyle('footer', parent=styles['Normal'],
                                   alignment=TA_CENTER, fontSize=8,
                                   textColor=colors.grey)
    elements.append(Paragraph('Thank you for your business!', footer_style))
    elements.append(Paragraph('Manamperi Rice Mill — Quality Rice Since Generations',
                              footer_style))

    doc.build(elements)

    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="invoice_{sale.invoice_number}.pdf"'
    )
    return response


# ======================================================================
# EXCEL REPORT EXPORT
# ======================================================================
def generate_excel_report(title, headers, rows):
    """
    Generate a styled Excel report.
    Returns an HttpResponse with the workbook attached.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Sheet name max 31 chars

    # Header style
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E',
                               fill_type='solid')
    header_align = Alignment(horizontal='center')

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, Decimal):
                cell.number_format = '#,##0.00'

    # Auto-size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_length, 30)

    buf = io.BytesIO()
    wb.save(buf)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{title.replace(" ", "_")}_{datetime.now():%Y%m%d}.xlsx"'
    )
    return response
